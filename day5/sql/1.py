import sqlite3

# username, passwd, url, dbname
objcon = sqlite3.connect('chinook.db')
objcur = objcon.cursor()

query = """SELECT * FROM employees"""

objcur.execute(query)

data = objcur.fetchall()

##################
from openpyxl import Workbook
wba = Workbook()
wsa = wba.active

wsa.cell(row=1, column=1).value = "First Name"
wsa.cell(row=1, column=2).value = "Second Name"
wsa.cell(row=1, column=3).value = "email ID"
for i in range(len(data)):
    wsa.cell(row=i+2, column=2).value = data[i][1]
    wsa.cell(row=i+2, column=1).value = data[i][2]
    wsa.cell(row=i+2, column=3).value = data[i][-1]

wba.save("data.xlsx")
