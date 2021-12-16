from openpyxl import load_workbook

wba = load_workbook("data1.xlsx")

#print(wba.sheetnames)

wsa = wba['monday']

print(wsa['C6'].value)
