from openpyxl import load_workbook

wba = load_workbook("data1.xlsx")

#print(wba.sheetnames)

wsa = wba['monday']

dicta = {}
for r in range(2, wsa.max_row+1):
    if wsa.cell(row=r, column=6).value > 100:
        dicta[(wsa.cell(row=r, column=3).value)] = wsa.cell(row=r, column=6).value

print(dicta)
