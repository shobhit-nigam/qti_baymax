from openpyxl import load_workbook

wba = load_workbook("data1.xlsx")

#print(wba.sheetnames)
objf = open("data1.txt", "w")

wsa = wba['monday']

for r in range(2, wsa.max_row+1):
    if wsa.cell(row=r, column=3).value == 'Carlos Soltero':
        for c in range(1, wsa.max_column+1):
            #print(wsa.cell(row=r, column=c).value, end=';')
            objf.write(str(wsa.cell(row=r, column=c).value)+'\t')
        objf.write("\n")

objf.close()
