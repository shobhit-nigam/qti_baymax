from openpyxl import load_workbook

wba = load_workbook("data1.xlsx")

#print(wba.sheetnames)
objf = open("data1.txt", "w")

wsa = wba['monday']

objc = wsa.cell(row=10, column=3).comment
print("author =",objc.author)
print("text =",objc.text)
