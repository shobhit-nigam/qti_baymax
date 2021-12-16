from openpyxl import load_workbook

wba = load_workbook("data1.xlsx")

#print(wba.sheetnames)

wsa = wba['monday']

cell_a = wsa.cell(row=5, column=4)
print(type(cell_a.value))
