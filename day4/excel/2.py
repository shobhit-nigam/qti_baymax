from openpyxl import load_workbook

wba = load_workbook("data1.xlsx")

#print(wba.sheetnames)

wsa = wba['monday']

cell_a = wsa['C6']

print(wsa.max_row)
print(wsa.max_column)
