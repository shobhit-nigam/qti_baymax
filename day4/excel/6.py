from openpyxl import load_workbook

wba = load_workbook("data1.xlsx")

#print(wba.sheetnames)

wsa = wba['monday']

lista= []
for r in range(2, wsa.max_row+1):
    if wsa.cell(row=r, column=6).value > 100:
        lista.append(wsa.cell(row=r, column=3).value)

print(lista)
