from openpyxl import load_workbook

wba = load_workbook("data1.xlsx")

#print(wba.sheetnames)

wsa = wba['monday']

for r in range(2, wsa.max_row+1):
    print(wsa.cell(row=r, column=6).value)     
