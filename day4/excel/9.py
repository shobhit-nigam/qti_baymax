from openpyxl import load_workbook

wba = load_workbook("data1.xlsx")

#print(wba.sheetnames)
objf = open("data1.txt", "w")

wsa = wba['monday']

for r in wsa.iter_rows(min_row=2, max_row=11, max_col=3):
    for c in r:
        print(c.value)
