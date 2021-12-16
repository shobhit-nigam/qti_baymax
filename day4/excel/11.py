from openpyxl import Workbook
import datetime
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.drawing.image import Image

wba = Workbook()

wsa = wba.active
wsa.title = "october"

wsa.cell(row=2, column=2).value = "hello"
wsa['B3'].value = "hi"
wsa['B4'] = datetime.datetime(2021, 12, 15)

lista = [11, 13, 15, 17, 16, 14, 12, 10, 13, 15]
wsb = wba.create_sheet("november")

for i in range(0, 10):
    wsb.cell(row=i+1, column=1).value = lista[i]

values = Reference(wsb, min_col=1, max_col=1, min_row=1, max_row=10)
objchart = BarChart()
objchart.add_data(values)

wsb.add_chart(objchart, "D5")

wsb = wba.create_sheet("december")
objimg = Image('q.png')
wsb.add_image(objimg, 'C3')

wba.save("data2.xlsx")
